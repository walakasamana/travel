"""Backend tests SSOT Batch 6 — UNDO gabung destinasi + KOTA BENGKEL + ekspor.

Cakupan:
  (A) UNDO GABUNGAN destinasi: merge menyimpan merged_moved; unmerge memulihkan;
      unmerge dgn skipped (dokumen berpindah manual); re-unmerge → 400.
  (B) KOTA BENGKEL: create/update workshops.city divalidasi city_or_400; canonicalisasi.
  (C) master/cities: setiap baris punya used_by_workshops.
  (D) Rename kota (PATCH /master/cities/{id}) cascade ke workshops.
  (E) Ekspor /master/export: sheet 'Kota' berisi kolom 'Bengkel'.
  (F) Regresi tipis: rename destinasi masih jalan; POST /api/customers city ngawur tetap 400.

Kebersihan: nama uji berawalan 'Penjaga INV-' + phone '0800000xxx'.
"""
import os
from datetime import datetime, timezone

import pytest
import requests

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "").rstrip("/")
API = f"{BASE_URL}/api"

CREDS = {"owner": ("owner@demo.local", "demo12345")}
GUARD_PHONE_PREFIX = "0800000"
GUARD_NAME_PREFIX = "Penjaga INV-"


def _login(email, pwd):
    r = requests.post(f"{API}/auth/login", json={"email": email, "password": pwd}, timeout=15)
    assert r.status_code == 200, f"login {email}: {r.status_code} {r.text}"
    tok = r.json().get("access_token") or r.json().get("token")
    return {"Authorization": f"Bearer {tok}"}


@pytest.fixture(scope="module")
def owner_h():
    return _login(*CREDS["owner"])


@pytest.fixture(scope="module")
def mongo():
    from pymongo import MongoClient
    mongo_url = os.environ.get("MONGO_URL")
    db_name = os.environ.get("DB_NAME")
    if not mongo_url or not db_name:
        try:
            with open("/app/backend/.env") as f:
                for line in f:
                    line = line.strip()
                    if line.startswith("MONGO_URL="):
                        mongo_url = line.split("=", 1)[1].strip().strip('"').strip("'")
                    elif line.startswith("DB_NAME="):
                        db_name = line.split("=", 1)[1].strip().strip('"').strip("'")
        except FileNotFoundError:
            pass
    assert mongo_url and db_name, "MONGO_URL / DB_NAME tidak tersedia"
    return MongoClient(mongo_url)[db_name]


# ============ (A) UNDO GABUNGAN destinasi ============
class TestUnmergeDestination:
    SRC_ID = "dst_test_b6"
    SRC_NAME = f"{GUARD_NAME_PREFIX}Bromo Kembar TEST"
    LEAD_ID = "led_test_b6_unmerge"
    Q_ID = "quo_test_b6_unmerge"

    def _seed_all(self, mongo):
        mongo.destinations.update_one(
            {"id": self.SRC_ID},
            {"$set": {"id": self.SRC_ID, "name": self.SRC_NAME,
                      "slug": "penjaga-bromo-kembar-test-b6",
                      "status": "draft", "ops_active": True,
                      "created_at": datetime.now(timezone.utc).isoformat()},
             "$unset": {"merged_into": "", "merged_into_name": "", "merged_moved": ""}},
            upsert=True)
        mongo.leads.update_one(
            {"id": self.LEAD_ID},
            {"$set": {"id": self.LEAD_ID, "name": f"{GUARD_NAME_PREFIX}UnmergeLead",
                      "phone": f"{GUARD_PHONE_PREFIX}601",
                      "destination": self.SRC_NAME, "stage": "new",
                      "source": "test",
                      "created_at": datetime.now(timezone.utc).isoformat()}},
            upsert=True)
        mongo.quotations.update_one(
            {"id": self.Q_ID},
            {"$set": {"id": self.Q_ID, "quote_no": "Q-TEST-B6",
                      "customer_id": "cust_seed_x", "vehicle_id": "veh_seed_x",
                      "destination": self.SRC_NAME, "status": "draft",
                      "created_at": datetime.now(timezone.utc).isoformat()}},
            upsert=True)

    def _cleanup(self, mongo):
        mongo.destinations.delete_one({"id": self.SRC_ID})
        mongo.leads.delete_one({"id": self.LEAD_ID})
        mongo.quotations.delete_one({"id": self.Q_ID})

    def test_full_merge_unmerge_flow(self, owner_h, mongo):
        self._seed_all(mongo)
        try:
            tgt = mongo.destinations.find_one({"name": "Gunung Bromo"}, {"_id": 0, "id": 1})
            assert tgt, "seed Gunung Bromo tidak ada"

            # Merge
            r = requests.post(f"{API}/master/destinations/{self.SRC_ID}/merge",
                              json={"target_id": tgt["id"]}, headers=owner_h, timeout=20)
            assert r.status_code == 200, r.text
            assert r.json().get("merged") is True

            # merged_moved berisi lead + quotation
            src = mongo.destinations.find_one({"id": self.SRC_ID}, {"_id": 0})
            moved = src.get("merged_moved") or {}
            assert self.LEAD_ID in (moved.get("leads") or []), f"leads not tracked: {moved}"
            assert self.Q_ID in (moved.get("quotations") or []), f"quotations not tracked: {moved}"
            assert src.get("ops_active") is False
            assert src.get("merged_into") == tgt["id"]

            # Verify dokumen dipindah ke Gunung Bromo
            assert mongo.leads.find_one({"id": self.LEAD_ID}).get("destination") == "Gunung Bromo"
            assert mongo.quotations.find_one({"id": self.Q_ID}).get("destination") == "Gunung Bromo"

            # UNMERGE
            r = requests.post(f"{API}/master/destinations/{self.SRC_ID}/unmerge",
                              headers=owner_h, timeout=20)
            assert r.status_code == 200, r.text
            body = r.json()
            assert body.get("unmerged") is True
            assert body["restored"]["leads"] == 1, body
            assert body["restored"]["quotations"] == 1, body
            assert body["skipped"] == 0, body

            # dokumen kembali ke nama sumber
            assert mongo.leads.find_one({"id": self.LEAD_ID}).get("destination") == self.SRC_NAME
            assert mongo.quotations.find_one({"id": self.Q_ID}).get("destination") == self.SRC_NAME
            src2 = mongo.destinations.find_one({"id": self.SRC_ID}, {"_id": 0})
            assert src2.get("ops_active") is True
            assert "merged_into" not in src2 or src2.get("merged_into") in (None, "")

            # Re-unmerge → 400
            r = requests.post(f"{API}/master/destinations/{self.SRC_ID}/unmerge",
                              headers=owner_h, timeout=15)
            assert r.status_code == 400, r.text
            assert "gabung" in (r.json().get("detail") or "").lower()
        finally:
            self._cleanup(mongo)

    def test_unmerge_with_skipped(self, owner_h, mongo):
        """Salah satu dokumen diubah manual ke destinasi lain sesudah merge → skipped."""
        self._seed_all(mongo)
        try:
            tgt = mongo.destinations.find_one({"name": "Gunung Bromo"}, {"_id": 0, "id": 1})
            # Merge
            r = requests.post(f"{API}/master/destinations/{self.SRC_ID}/merge",
                              json={"target_id": tgt["id"]}, headers=owner_h, timeout=20)
            assert r.status_code == 200, r.text

            # Ubah manual lead ke destinasi lain 'Bali'
            mongo.leads.update_one({"id": self.LEAD_ID}, {"$set": {"destination": "Bali"}})

            # Unmerge → lead diskip, quotation kembali
            r = requests.post(f"{API}/master/destinations/{self.SRC_ID}/unmerge",
                              headers=owner_h, timeout=20)
            assert r.status_code == 200, r.text
            body = r.json()
            assert body["restored"]["leads"] == 0, body
            assert body["restored"]["quotations"] == 1, body
            assert body["skipped"] == 1, body

            assert mongo.leads.find_one({"id": self.LEAD_ID}).get("destination") == "Bali"
            assert mongo.quotations.find_one({"id": self.Q_ID}).get("destination") == self.SRC_NAME
        finally:
            self._cleanup(mongo)


# ============ (B) KOTA BENGKEL ============
class TestWorkshopCity:
    created_ids = []

    def test_create_bad_city_400(self, owner_h):
        payload = {"name": f"{GUARD_NAME_PREFIX}WSHBadCity",
                   "phone": f"{GUARD_PHONE_PREFIX}701",
                   "city": "KotaNgawurB6"}
        r = requests.post(f"{API}/workshops", json=payload, headers=owner_h, timeout=15)
        assert r.status_code == 400, r.text
        detail = (r.json().get("detail") or "").lower()
        assert "kota" in detail or "master" in detail, detail

    def test_create_city_canonicalized(self, owner_h):
        payload = {"name": f"{GUARD_NAME_PREFIX}WSHCanon",
                   "phone": f"{GUARD_PHONE_PREFIX}702",
                   "city": "bandung"}
        r = requests.post(f"{API}/workshops", json=payload, headers=owner_h, timeout=15)
        assert r.status_code == 200, r.text
        assert r.json().get("city") == "Bandung"
        TestWorkshopCity.created_ids.append(r.json().get("id"))

    def test_patch_bad_city_400(self, owner_h):
        # buat bengkel valid, lalu patch ke kota ngawur
        payload = {"name": f"{GUARD_NAME_PREFIX}WSHPatch",
                   "phone": f"{GUARD_PHONE_PREFIX}703",
                   "city": "Bandung"}
        r = requests.post(f"{API}/workshops", json=payload, headers=owner_h, timeout=15)
        assert r.status_code == 200, r.text
        wid = r.json().get("id")
        TestWorkshopCity.created_ids.append(wid)
        r2 = requests.patch(f"{API}/workshops/{wid}",
                            json={"city": "KotaNgawurB6"}, headers=owner_h, timeout=15)
        assert r2.status_code == 400, r2.text

    def test_zzz_cleanup_workshops(self, owner_h):
        for wid in TestWorkshopCity.created_ids:
            requests.delete(f"{API}/workshops/{wid}", headers=owner_h, timeout=15)


# ============ (C) master/cities used_by_workshops ============
class TestMasterCitiesUsedByWorkshops:
    def test_rows_have_used_by_workshops(self, owner_h):
        r = requests.get(f"{API}/master/cities", headers=owner_h, timeout=15)
        assert r.status_code == 200, r.text
        rows = r.json()
        assert isinstance(rows, list) and rows
        for c in rows:
            assert "used_by_workshops" in c, f"missing used_by_workshops in {c}"
            assert isinstance(c["used_by_workshops"], int)


# ============ (D) Rename kota cascade ke workshops ============
class TestCityRenameCascadeWorkshops:
    def test_cascade(self, owner_h, mongo):
        # Seed kota uji + workshop menempel kota tsb
        city_name = f"{GUARD_NAME_PREFIX}Ciamis Uji B6"
        rc = requests.post(f"{API}/cities", json={"name": city_name},
                           headers=owner_h, timeout=15)
        assert rc.status_code == 200, rc.text
        cid = rc.json()["id"]

        rw = requests.post(f"{API}/workshops",
                           json={"name": f"{GUARD_NAME_PREFIX}WSHCascade",
                                 "phone": f"{GUARD_PHONE_PREFIX}704",
                                 "city": city_name}, headers=owner_h, timeout=15)
        assert rw.status_code == 200, rw.text
        wid = rw.json()["id"]

        try:
            new_name = f"{GUARD_NAME_PREFIX}Ciamis Baru B6"
            rp = requests.patch(f"{API}/master/cities/{cid}",
                                json={"name": new_name}, headers=owner_h, timeout=15)
            assert rp.status_code == 200, rp.text
            cascade = rp.json().get("cascade") or {}
            assert cascade.get("workshops", 0) >= 1, cascade

            # verifikasi workshop punya kota baru
            wg = requests.get(f"{API}/workshops", headers=owner_h, timeout=15).json()
            w = [x for x in wg if x.get("id") == wid][0]
            assert w.get("city") == new_name
        finally:
            requests.delete(f"{API}/workshops/{wid}", headers=owner_h, timeout=15)
            # soft-remove kota uji dari master (tak ada DELETE endpoint → hard delete via mongo)
            mongo.cities.delete_one({"id": cid})


# ============ (E) Ekspor sheet Kota berkolom Bengkel ============
class TestExportKotaSheetBengkel:
    def test_sheet_kota_has_bengkel_column(self, owner_h):
        r = requests.get(f"{API}/master/export", headers=owner_h, timeout=30)
        assert r.status_code == 200, r.text
        from io import BytesIO
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(r.content), read_only=True)
        assert "Kota" in wb.sheetnames
        ws = wb["Kota"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        assert "Bengkel" in headers, f"headers Kota: {headers}"


# ============ (F) Regresi tipis ============
class TestRegressionThin:
    def test_rename_destination_still_works(self, owner_h, mongo):
        """Rename destinasi (bukan merge) masih menerima cascade preview minimal 200."""
        # pilih destinasi apa saja yg tak digabung
        d = mongo.destinations.find_one(
            {"deleted": {"$ne": True}, "merged_into": {"$in": [None, ""]}, "name": "Bali"},
            {"_id": 0, "id": 1, "name": 1})
        if not d:
            d = mongo.destinations.find_one({"deleted": {"$ne": True}}, {"_id": 0, "id": 1, "name": 1})
        assert d, "butuh minimal 1 destinasi utk regresi rename"
        original = d["name"]
        # rename ke nama yg sama → seharusnya 200 tanpa cascade (no-op)
        r = requests.patch(f"{API}/master/destinations/{d['id']}",
                           json={"name": original}, headers=owner_h, timeout=15)
        assert r.status_code == 200, r.text

    def test_customer_bad_city_still_400(self, owner_h):
        payload = {"name": f"{GUARD_NAME_PREFIX}RegCustBad",
                   "phone": f"{GUARD_PHONE_PREFIX}901",
                   "type": "individual", "city": "KotaAmburadulB6"}
        r = requests.post(f"{API}/customers", json=payload, headers=owner_h, timeout=15)
        assert r.status_code == 400, r.text
