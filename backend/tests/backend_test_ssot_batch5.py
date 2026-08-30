"""Backend tests SSOT Batch 5 — Master KOTA + Ekspor Excel + Gabung Duplikat destinasi.

Cakupan:
  (A) KOTA: customers/partners city_or_400 create+update; POST /api/cities idempotent; GET /api/cities.
  (B) vehicle_type_normalize di /api/public/landing/{slug}/lead.
  (C) GET /api/master/export owner=xlsx 3 sheet; RBAC marketing_admin & driver → 403.
  (D) Gabung destinasi: sukses cascade, error re-merge/self/target nonaktif; cleanup Mongo.
  (E) Regresi tipis: /api/bookings destinasi ngawur tetap 400.

Kebersihan: nama uji berawalan 'Penjaga INV-' + phone '0800000xxx' agar tercakup purge guard.
"""
import os
import time
from datetime import datetime, timedelta, timezone

import pytest
import requests

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "").rstrip("/")
API = f"{BASE_URL}/api"

CREDS = {
    "owner": ("owner@demo.local", "demo12345"),
    "marketing_admin": ("marketing@demo.local", "demo12345"),
    "driver": ("driver@demo.local", "demo12345"),
}
GUARD_PHONE_PREFIX = "0800000"
GUARD_NAME_PREFIX = "Penjaga INV-"


def _login(email, pwd):
    r = requests.post(f"{API}/auth/login", json={"email": email, "password": pwd}, timeout=15)
    assert r.status_code == 200, f"login {email}: {r.status_code} {r.text}"
    tok = r.json().get("access_token") or r.json().get("token")
    return {"Authorization": f"Bearer {tok}"}


@pytest.fixture(scope="module")
def owner_h():
    return _login(*CREDS["owner"])


@pytest.fixture(scope="module")
def marketing_h():
    return _login(*CREDS["marketing_admin"])


@pytest.fixture(scope="module")
def driver_h():
    return _login(*CREDS["driver"])


def _future_iso(hours=48):
    return (datetime.now(timezone.utc) + timedelta(hours=hours)).replace(microsecond=0).isoformat()


# ============ (A) KOTA ============
class TestCityMaster:
    def test_list_cities_active(self, owner_h):
        r = requests.get(f"{API}/cities", headers=owner_h, timeout=15)
        assert r.status_code == 200, r.text
        rows = r.json()
        assert isinstance(rows, list) and len(rows) > 0
        names = [c.get("name") for c in rows]
        assert "Bandung" in names, f"Bandung missing from seeded cities: {names}"

    def test_post_city_idempotent(self, owner_h):
        payload = {"name": f"{GUARD_NAME_PREFIX}Kota Uji Penjaga INV-REF-02"}
        r1 = requests.post(f"{API}/cities", json=payload, headers=owner_h, timeout=15)
        assert r1.status_code == 200, r1.text
        id1 = r1.json().get("id")
        assert id1
        r2 = requests.post(f"{API}/cities", json=payload, headers=owner_h, timeout=15)
        assert r2.status_code == 200, r2.text
        assert r2.json().get("id") == id1, "quick-add city not idempotent"

    def test_create_customer_bad_city_400(self, owner_h):
        payload = {
            "name": f"{GUARD_NAME_PREFIX}CustBadCity",
            "phone": f"{GUARD_PHONE_PREFIX}201",
            "type": "individual",
            "city": "KotaNgawurXYZ",
        }
        r = requests.post(f"{API}/customers", json=payload, headers=owner_h, timeout=15)
        assert r.status_code == 400, r.text
        detail = (r.json().get("detail") or "").lower()
        assert "kota" in detail or "master" in detail, f"expected reason mention kota/master: {detail}"

    def test_create_customer_city_case_normalized(self, owner_h):
        payload = {
            "name": f"{GUARD_NAME_PREFIX}CustCanon",
            "phone": f"{GUARD_PHONE_PREFIX}202",
            "type": "individual",
            "city": "bandung",
        }
        r = requests.post(f"{API}/customers", json=payload, headers=owner_h, timeout=15)
        assert r.status_code == 200, r.text
        assert r.json().get("city") == "Bandung"

    def test_create_customer_empty_city_ok(self, owner_h):
        payload = {
            "name": f"{GUARD_NAME_PREFIX}CustNoCity",
            "phone": f"{GUARD_PHONE_PREFIX}203",
            "type": "individual",
            "city": "",
        }
        r = requests.post(f"{API}/customers", json=payload, headers=owner_h, timeout=15)
        assert r.status_code == 200, r.text
        assert (r.json().get("city") or "") == ""

    def test_patch_customer_bad_city_400(self, owner_h):
        payload = {
            "name": f"{GUARD_NAME_PREFIX}CustPatch",
            "phone": f"{GUARD_PHONE_PREFIX}204",
            "type": "individual",
            "city": "Bandung",
        }
        r = requests.post(f"{API}/customers", json=payload, headers=owner_h, timeout=15)
        assert r.status_code == 200, r.text
        cid = r.json().get("id")
        r2 = requests.patch(f"{API}/customers/{cid}", json={"city": "KotaNgawurXYZ"}, headers=owner_h, timeout=15)
        assert r2.status_code == 400, r2.text

    def test_create_partner_bad_city_400(self, owner_h):
        payload = {
            "name": f"{GUARD_NAME_PREFIX}PartnerBadCity",
            "phone": f"{GUARD_PHONE_PREFIX}301",
            "city": "KotaNgawurXYZ",
        }
        r = requests.post(f"{API}/partners", json=payload, headers=owner_h, timeout=15)
        assert r.status_code == 400, r.text

    def test_create_partner_city_normalized(self, owner_h):
        payload = {
            "name": f"{GUARD_NAME_PREFIX}PartnerCanon",
            "phone": f"{GUARD_PHONE_PREFIX}302",
            "city": "jakarta",
        }
        r = requests.post(f"{API}/partners", json=payload, headers=owner_h, timeout=15)
        assert r.status_code == 200, r.text
        assert r.json().get("city") == "Jakarta"

    def test_patch_partner_bad_city_400(self, owner_h):
        payload = {
            "name": f"{GUARD_NAME_PREFIX}PartnerPatch",
            "phone": f"{GUARD_PHONE_PREFIX}303",
            "city": "Jakarta",
        }
        r = requests.post(f"{API}/partners", json=payload, headers=owner_h, timeout=15)
        assert r.status_code == 200, r.text
        pid = r.json().get("id")
        r2 = requests.patch(f"{API}/partners/{pid}", json={"city": "KotaNgawurXYZ"}, headers=owner_h, timeout=15)
        assert r2.status_code == 400, r2.text


# ============ (B) vehicle_type_normalize di landing lead ============
class TestLandingVehicleTypeSoftNormalize:
    def test_label_normalized_to_key(self, owner_h):
        # Slug seed: sewa-hiace-jakarta
        payload = {
            "name": f"{GUARD_NAME_PREFIX}VTCanon",
            "phone": f"{GUARD_PHONE_PREFIX}401",
            "vehicle_type": "Hiace Premio",
            "marketing_consent": True,
            "message": "TEST batch 5 vehicle_type label",
        }
        time.sleep(1)
        r = requests.post(f"{API}/public/landing/sewa-hiace-jakarta/lead", json=payload, timeout=20)
        assert r.status_code == 200, r.text
        lid = r.json().get("id") or r.json().get("lead_id")
        assert lid
        # Verify via owner-scoped leads list
        time.sleep(0.4)
        rl = requests.get(f"{API}/leads", headers=owner_h, params={"limit": 200}, timeout=15)
        assert rl.status_code == 200
        found = [x for x in rl.json() if x.get("id") == lid]
        assert found, f"lead {lid} not found"
        assert found[0].get("vehicle_type") == "hiace_premio", f"got {found[0].get('vehicle_type')}"

    def test_unknown_vehicle_type_soft_accepted(self, owner_h):
        payload = {
            "name": f"{GUARD_NAME_PREFIX}VTUnknown",
            "phone": f"{GUARD_PHONE_PREFIX}402",
            "vehicle_type": "mobil aneh",
            "marketing_consent": True,
            "message": "TEST batch 5 vehicle_type unknown",
        }
        time.sleep(1)
        r = requests.post(f"{API}/public/landing/sewa-hiace-jakarta/lead", json=payload, timeout=20)
        assert r.status_code == 200, r.text
        lid = r.json().get("id") or r.json().get("lead_id")
        time.sleep(0.4)
        rl = requests.get(f"{API}/leads", headers=owner_h, params={"limit": 200}, timeout=15)
        assert rl.status_code == 200
        found = [x for x in rl.json() if x.get("id") == lid]
        assert found
        assert found[0].get("vehicle_type") == "mobil aneh", f"soft normalize should keep original, got {found[0].get('vehicle_type')}"


# ============ (C) Ekspor Excel ============
class TestMasterExport:
    def test_export_owner_xlsx(self, owner_h):
        r = requests.get(f"{API}/master/export", headers=owner_h, timeout=30)
        assert r.status_code == 200, r.text
        ctype = r.headers.get("Content-Type", "")
        assert "spreadsheetml" in ctype, ctype
        content = r.content
        assert content[:2] == b"PK", "xlsx should be a zip (PK) file"
        # Verify sheets in-memory
        from io import BytesIO
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(content), read_only=True)
        assert set(wb.sheetnames) >= {"Titik Jemput", "Destinasi", "Kota"}, wb.sheetnames

    def test_export_marketing_forbidden(self, marketing_h):
        r = requests.get(f"{API}/master/export", headers=marketing_h, timeout=15)
        assert r.status_code == 403, r.text

    def test_export_driver_forbidden(self, driver_h):
        r = requests.get(f"{API}/master/export", headers=driver_h, timeout=15)
        assert r.status_code == 403, r.text


# ============ (D) Gabung destinasi ============
class TestDestinationMerge:
    """Menggunakan koneksi Mongo langsung utk siapkan/hapus artefak uji."""

    @pytest.fixture(scope="class")
    def mongo(self):
        from pymongo import MongoClient
        mongo_url = os.environ.get("MONGO_URL")
        db_name = os.environ.get("DB_NAME")
        if not mongo_url or not db_name:
            try:
                with open("/app/backend/.env") as f:
                    for line in f:
                        line = line.strip()
                        if line.startswith("MONGO_URL="):
                            mongo_url = line.split("=", 1)[1].strip().strip('"').strip("'")
                        elif line.startswith("DB_NAME="):
                            db_name = line.split("=", 1)[1].strip().strip('"').strip("'")
            except FileNotFoundError:
                pass
        assert mongo_url and db_name, "MONGO_URL / DB_NAME tidak tersedia"
        client = MongoClient(mongo_url)
        return client[db_name]

    @pytest.fixture(scope="class")
    def merge_setup(self, mongo):
        src_id = "dst_test_b5"
        src_name = f"{GUARD_NAME_PREFIX}Bromo Kembar TEST"
        mongo.destinations.update_one(
            {"id": src_id},
            {"$set": {"id": src_id, "name": src_name, "slug": "penjaga-bromo-kembar-test",
                      "status": "draft", "ops_active": True,
                      "created_at": datetime.now(timezone.utc).isoformat()}},
            upsert=True)
        tgt = mongo.destinations.find_one({"name": "Gunung Bromo"}, {"_id": 0, "id": 1})
        assert tgt, "seed 'Gunung Bromo' tidak ditemukan"
        lead_id = "led_test_b5_merge"
        mongo.leads.update_one(
            {"id": lead_id},
            {"$set": {"id": lead_id, "name": f"{GUARD_NAME_PREFIX}MergeLead",
                      "phone": f"{GUARD_PHONE_PREFIX}501",
                      "destination": src_name, "stage": "new",
                      "source": "test", "created_at": datetime.now(timezone.utc).isoformat()}},
            upsert=True)
        try:
            yield src_id, src_name, tgt["id"], lead_id
        finally:
            mongo.destinations.delete_one({"id": "dst_test_b5"})
            mongo.leads.delete_one({"id": "led_test_b5_merge"})

    def test_merge_self_400(self, owner_h, merge_setup):
        src_id, *_ = merge_setup
        r = requests.post(f"{API}/master/destinations/{src_id}/merge",
                          json={"target_id": src_id}, headers=owner_h, timeout=15)
        assert r.status_code == 400, r.text

    def test_merge_target_inactive_400(self, owner_h, merge_setup, mongo):
        src_id, *_ = merge_setup
        tgt = mongo.destinations.find_one({"name": "Jakarta"}, {"_id": 0, "id": 1, "ops_active": 1})
        assert tgt, "seed Jakarta tidak ditemukan"
        prev = tgt.get("ops_active", True)
        mongo.destinations.update_one({"id": tgt["id"]}, {"$set": {"ops_active": False}})
        try:
            r = requests.post(f"{API}/master/destinations/{src_id}/merge",
                              json={"target_id": tgt["id"]}, headers=owner_h, timeout=15)
            assert r.status_code == 400, r.text
        finally:
            mongo.destinations.update_one({"id": tgt["id"]},
                                          {"$set": {"ops_active": prev if prev is not None else True}})

    def test_merge_success_cascade_and_flags(self, owner_h, merge_setup, mongo):
        src_id, src_name, tgt_id, lead_id = merge_setup
        r = requests.post(f"{API}/master/destinations/{src_id}/merge",
                          json={"target_id": tgt_id}, headers=owner_h, timeout=20)
        assert r.status_code == 200, r.text
        body = r.json()
        assert body.get("merged") is True
        assert body["cascade"]["leads"] >= 1
        src_doc = mongo.destinations.find_one({"id": src_id}, {"_id": 0})
        assert src_doc.get("ops_active") is False
        assert src_doc.get("merged_into") == tgt_id
        assert src_doc.get("merged_into_name") == "Gunung Bromo"
        lead = mongo.leads.find_one({"id": lead_id}, {"_id": 0})
        assert lead.get("destination") == "Gunung Bromo"

    def test_merge_again_400(self, owner_h, merge_setup):
        src_id, _, tgt_id, _ = merge_setup
        r = requests.post(f"{API}/master/destinations/{src_id}/merge",
                          json={"target_id": tgt_id}, headers=owner_h, timeout=15)
        assert r.status_code == 400, r.text


# ============ (E) Regresi: booking ERP dgn destinasi ngawur → 400 ============
class TestBookingHardValidationRegression:
    def test_booking_bad_destination_still_400(self, owner_h):
        # Ambil customer & vehicle apa saja utk memenuhi field wajib.
        rc = requests.get(f"{API}/customers", headers=owner_h, params={"limit": 1}, timeout=15)
        assert rc.status_code == 200 and rc.json(), "butuh minimal 1 customer utk regresi"
        cust_id = rc.json()[0]["id"]
        rv = requests.get(f"{API}/vehicles", headers=owner_h, timeout=15)
        assert rv.status_code == 200 and rv.json(), "butuh minimal 1 vehicle utk regresi"
        veh_id = rv.json()[0]["id"]
        payload = {
            "customer_id": cust_id,
            "vehicle_id": veh_id,
            "origin": "Bandung",
            "destination": "KampungAntahBerantahXYZ",
            "start_datetime": _future_iso(48),
            "end_datetime": _future_iso(72),
            "pax": 2,
        }
        r = requests.post(f"{API}/bookings", json=payload, headers=owner_h, timeout=15)
        assert r.status_code == 400, r.text
        assert "destin" in (r.json().get("detail") or "").lower()
