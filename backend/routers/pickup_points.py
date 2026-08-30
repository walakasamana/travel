"""routers/pickup_points.py — Master Data referensi (INV-REF-02): TITIK JEMPUT + DESTINASI + KOTA.

`bookings.origin`, destinasi (booking/lead/penawaran), dan kota (customers/partners) bukan teks
bebas — nilainya wajib dari master di file ini / koleksi terkait. Halaman kelola: /app/masterdata
(section `masterdata`, owner + ops_admin). RENAME di sini CASCADE ke dokumen pemakai supaya nama
kanonik tidak pernah bercabang; NONAKTIF menyembunyikan dari selector & menolak pemakaian BARU
tanpa merusak data lama. Batch 5: master KOTA, GABUNG destinasi kembar, ekspor Excel.
"""
from io import BytesIO

from core_utils import new_id, now_iso, safe_doc
from db import get_db
from dependencies import require_section
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from schemas import (
    CityCreate,
    DestinationMergeRequest,
    MasterCityUpdate,
    MasterDestinationUpdate,
    PickupPointCreate,
    PickupPointUpdate,
)
from services.audit import record

router = APIRouter(prefix="/api", tags=["master-data"])
BOOKINGS = require_section("bookings")
MASTER = require_section("masterdata")
CITIES = require_section("customers")


def _clean(name: str) -> str:
    return str(name or "").strip()


async def _name_taken(coll, name: str, exclude_id: str = "") -> bool:
    async for d in coll.find({"deleted": {"$ne": True}}, {"_id": 0, "id": 1, "name": 1}):
        if d.get("id") != exclude_id and _clean(d.get("name")).lower() == name.lower():
            return True
    return False


@router.get("/pickup-points")
async def list_pickup_points(user=Depends(BOOKINGS)):
    """Utk selector form (hanya yang AKTIF). Kelola lengkap: GET /master/pickup-points."""
    rows = await get_db().pickup_points.find(
        {"deleted": {"$ne": True}, "active": {"$ne": False}},
        {"_id": 0}).sort("name", 1).to_list(500)
    return safe_doc(rows)


@router.post("/pickup-points")
async def create_pickup_point(body: PickupPointCreate, user=Depends(BOOKINGS)):
    """Quick-add master. Idempotent: nama yang sudah ada (case-insensitive) dikembalikan, bukan duplikat."""
    db = get_db()
    name = body.name.strip()
    async for p in db.pickup_points.find({"deleted": {"$ne": True}}, {"_id": 0}):
        if _clean(p.get("name")).lower() == name.lower():
            if p.get("active") is False:
                await db.pickup_points.update_one({"id": p["id"]}, {"$set": {"active": True}})
                p["active"] = True
            return safe_doc(p)
    doc = {"id": new_id("pkp"), "name": name, "active": True, "created_at": now_iso()}
    await db.pickup_points.insert_one(dict(doc))
    await record(db, actor=user, action="create", entity_type="pickup_point",
                 entity_id=doc["id"], summary=f"Titik jemput baru: {name}")
    return doc


@router.get("/master/pickup-points")
async def master_pickup_points(user=Depends(MASTER)):
    """Kelola master: semua baris (termasuk nonaktif) + jumlah pemakaian di booking."""
    db = get_db()
    rows = await db.pickup_points.find({"deleted": {"$ne": True}}, {"_id": 0}).sort("name", 1).to_list(500)
    out = []
    for r in rows:
        used = await db.bookings.count_documents({"origin": r.get("name")})
        moved = r.pop("merged_moved", None) or {}
        out.append({**r, "active": r.get("active") is not False, "used_by_bookings": used,
                    "merged_moved_count": sum(len(v or []) for v in moved.values())})
    return safe_doc(out)


@router.patch("/master/pickup-points/{point_id}")
async def update_pickup_point(point_id: str, body: PickupPointUpdate, user=Depends(MASTER)):
    """Rename (CASCADE ke bookings.origin) dan/atau aktif/nonaktif."""
    db = get_db()
    point = await db.pickup_points.find_one({"id": point_id, "deleted": {"$ne": True}}, {"_id": 0})
    if not point:
        raise HTTPException(status_code=404, detail="Titik jemput tidak ditemukan")
    updates, cascaded = {}, 0
    new_name = _clean(body.name) if body.name is not None else ""
    old_name = _clean(point.get("name"))
    if new_name and new_name != old_name:
        if await _name_taken(db.pickup_points, new_name, exclude_id=point_id):
            raise HTTPException(status_code=400, detail=f"Nama '{new_name}' sudah dipakai baris master lain")
        updates["name"] = new_name
        res = await db.bookings.update_many({"origin": old_name}, {"$set": {"origin": new_name}})
        cascaded = res.modified_count
    if body.active is not None:
        updates["active"] = bool(body.active)
    if not updates:
        return {**point, "cascaded_bookings": 0}
    await db.pickup_points.update_one({"id": point_id}, {"$set": updates})
    await record(db, actor=user, action="update", entity_type="pickup_point", entity_id=point_id,
                 summary=f"Master titik jemput: {old_name} → {updates.get('name', old_name)}"
                         f"{' (nonaktif)' if updates.get('active') is False else ''}"
                         f" · cascade {cascaded} booking")
    return {**point, **updates, "cascaded_bookings": cascaded}


@router.post("/master/pickup-points/{point_id}/merge")
async def merge_master_pickup_point(point_id: str, body: DestinationMergeRequest, user=Depends(MASTER)):
    """GABUNG titik jemput kembar (batch 7): booking SUMBER pindah memakai nama TARGET,
    lalu sumber dinonaktifkan + ditandai `merged_into` (tidak ada data yang dihapus)."""
    db = get_db()
    if point_id == body.target_id:
        raise HTTPException(status_code=400, detail="Sumber dan target gabung tidak boleh sama")
    source = await db.pickup_points.find_one({"id": point_id, "deleted": {"$ne": True}}, {"_id": 0})
    target = await db.pickup_points.find_one({"id": body.target_id, "deleted": {"$ne": True}}, {"_id": 0})
    if not source or not target:
        raise HTTPException(status_code=404, detail="Titik jemput sumber/target tidak ditemukan")
    if source.get("merged_into"):
        raise HTTPException(status_code=400, detail="Titik jemput ini sudah pernah digabung")
    if target.get("merged_into") or target.get("active") is False:
        raise HTTPException(status_code=400, detail="Target gabung harus titik jemput AKTIF (bukan hasil gabungan/nonaktif)")
    src_name, tgt_name = _clean(source.get("name")), _clean(target.get("name"))
    ids = [d["id"] async for d in db.bookings.find(
        {"origin": src_name}, {"_id": 0, "id": 1}) if d.get("id")]
    cascaded = 0
    if ids:
        res = await db.bookings.update_many({"origin": src_name}, {"$set": {"origin": tgt_name}})
        cascaded = res.modified_count
    await db.pickup_points.update_one({"id": point_id}, {"$set": {
        "active": False, "merged_into": target["id"], "merged_into_name": tgt_name,
        "merged_moved": {"bookings": ids}, "updated_at": now_iso()}})
    await record(db, actor=user, action="update", entity_type="pickup_point", entity_id=point_id,
                 summary=f"GABUNG titik jemput: '{src_name}' → '{tgt_name}' "
                         f"· cascade {cascaded} booking")
    return {"merged": True, "source": src_name, "target": tgt_name,
            "cascade": {"bookings": cascaded}}


@router.post("/master/pickup-points/{point_id}/unmerge")
async def unmerge_master_pickup_point(point_id: str, user=Depends(MASTER)):
    """BATALKAN gabungan titik jemput: booking yang IKUT PINDAH (tercatat `merged_moved`)
    dikembalikan — hanya bila origin-nya masih nama target; sisanya `skipped`."""
    db = get_db()
    source = await db.pickup_points.find_one({"id": point_id, "deleted": {"$ne": True}}, {"_id": 0})
    if not source:
        raise HTTPException(status_code=404, detail="Titik jemput tidak ditemukan")
    if not source.get("merged_into"):
        raise HTTPException(status_code=400, detail="Titik jemput ini tidak sedang dalam status gabungan")
    src_name = _clean(source.get("name"))
    target = await db.pickup_points.find_one({"id": source["merged_into"]}, {"_id": 0})
    tgt_name = _clean((target or {}).get("name")) or _clean(source.get("merged_into_name"))
    moved = source.get("merged_moved") or {}
    restored, skipped = 0, 0
    for doc_id in (moved.get("bookings") or []):
        res = await db.bookings.update_one(
            {"id": doc_id, "origin": tgt_name}, {"$set": {"origin": src_name}})
        if res.modified_count:
            restored += 1
        else:
            skipped += 1
    await db.pickup_points.update_one({"id": point_id}, {
        "$set": {"active": True, "updated_at": now_iso()},
        "$unset": {"merged_into": "", "merged_into_name": "", "merged_moved": ""}})
    await record(db, actor=user, action="update", entity_type="pickup_point", entity_id=point_id,
                 summary=f"BATAL GABUNG titik jemput: '{src_name}' dipulihkan dari '{tgt_name}' "
                         f"· {restored} booking kembali, {skipped} dilewati")
    return {"unmerged": True, "source": src_name, "from_target": tgt_name,
            "restored": {"bookings": restored}, "skipped": skipped}


@router.get("/master/destinations")
async def master_destinations(user=Depends(MASTER)):
    """Kelola master destinasi dari sisi OPS: nama, status ops (aktif utk selector), pemakaian."""
    db = get_db()
    rows = await db.destinations.find(
        {"deleted": {"$ne": True}},
        {"_id": 0, "id": 1, "name": 1, "slug": 1, "status": 1, "source": 1, "ops_active": 1,
         "merged_into": 1, "merged_into_name": 1, "merged_moved": 1},
    ).sort("name", 1).to_list(500)
    out = []
    for r in rows:
        nm = r.get("name")
        used_b = await db.bookings.count_documents({"destination": nm})
        used_l = await db.leads.count_documents({"destination": nm})
        used_q = await db.quotations.count_documents({"destination": nm})
        moved = r.pop("merged_moved", None) or {}
        out.append({**r, "ops_active": r.get("ops_active") is not False,
                    "used_by_bookings": used_b, "used_by_leads": used_l,
                    "used_by_quotations": used_q,
                    "merged_moved_count": sum(len(v or []) for v in moved.values())})
    return safe_doc(out)


@router.patch("/master/destinations/{dest_id}")
async def update_master_destination(dest_id: str, body: MasterDestinationUpdate, user=Depends(MASTER)):
    """Rename destinasi (CASCADE ke bookings/leads/quotations yang memakai nama lama) +
    toggle `ops_active` (nonaktif = hilang dari selector & ditolak utk pemakaian baru;
    halaman web publik TIDAK berubah — slug tetap)."""
    db = get_db()
    dest = await db.destinations.find_one({"id": dest_id, "deleted": {"$ne": True}}, {"_id": 0})
    if not dest:
        raise HTTPException(status_code=404, detail="Destinasi tidak ditemukan")
    updates, cascade = {}, {"bookings": 0, "leads": 0, "quotations": 0}
    new_name = _clean(body.name) if body.name is not None else ""
    old_name = _clean(dest.get("name"))
    if new_name and new_name != old_name:
        if await _name_taken(db.destinations, new_name, exclude_id=dest_id):
            raise HTTPException(status_code=400, detail=f"Nama '{new_name}' sudah dipakai destinasi lain")
        updates["name"] = new_name
        for coll in ("bookings", "leads", "quotations"):
            res = await db[coll].update_many({"destination": old_name},
                                             {"$set": {"destination": new_name}})
            cascade[coll] = res.modified_count
    if body.ops_active is not None:
        updates["ops_active"] = bool(body.ops_active)
    if not updates:
        return {**dest, "cascade": cascade}
    updates["updated_at"] = now_iso()
    await db.destinations.update_one({"id": dest_id}, {"$set": updates})
    await record(db, actor=user, action="update", entity_type="destination", entity_id=dest_id,
                 summary=f"Master destinasi: {old_name} → {updates.get('name', old_name)}"
                         f"{' (ops nonaktif)' if updates.get('ops_active') is False else ''}"
                         f" · cascade {sum(cascade.values())} dokumen")
    return {**dest, **updates, "cascade": cascade}


@router.post("/master/destinations/{dest_id}/merge")
async def merge_master_destination(dest_id: str, body: DestinationMergeRequest, user=Depends(MASTER)):
    """GABUNG destinasi kembar (batch 5): seluruh booking/lead/penawaran SUMBER pindah ke
    nama TARGET, lalu sumber di-nonaktifkan + ditandai `merged_into` (riwayat menyatu,
    tidak ada data yang dihapus; halaman web publik target tidak berubah)."""
    db = get_db()
    if dest_id == body.target_id:
        raise HTTPException(status_code=400, detail="Sumber dan target gabung tidak boleh sama")
    source = await db.destinations.find_one({"id": dest_id, "deleted": {"$ne": True}}, {"_id": 0})
    target = await db.destinations.find_one({"id": body.target_id, "deleted": {"$ne": True}}, {"_id": 0})
    if not source or not target:
        raise HTTPException(status_code=404, detail="Destinasi sumber/target tidak ditemukan")
    if source.get("merged_into"):
        raise HTTPException(status_code=400, detail="Destinasi ini sudah pernah digabung")
    if target.get("merged_into") or target.get("ops_active") is False:
        raise HTTPException(status_code=400, detail="Target gabung harus destinasi AKTIF (bukan hasil gabungan/nonaktif)")
    src_name, tgt_name = _clean(source.get("name")), _clean(target.get("name"))
    cascade, moved = {"bookings": 0, "leads": 0, "quotations": 0}, {}
    for coll in ("bookings", "leads", "quotations"):
        ids = [d["id"] async for d in db[coll].find(
            {"destination": src_name}, {"_id": 0, "id": 1}) if d.get("id")]
        if ids:
            res = await db[coll].update_many({"destination": src_name},
                                             {"$set": {"destination": tgt_name}})
            cascade[coll] = res.modified_count
            moved[coll] = ids
    await db.destinations.update_one({"id": dest_id}, {"$set": {
        "ops_active": False, "merged_into": target["id"], "merged_into_name": tgt_name,
        "merged_moved": moved, "updated_at": now_iso()}})
    await record(db, actor=user, action="update", entity_type="destination", entity_id=dest_id,
                 summary=f"GABUNG destinasi: '{src_name}' → '{tgt_name}' "
                         f"· cascade {sum(cascade.values())} dokumen")
    return {"merged": True, "source": src_name, "target": tgt_name, "cascade": cascade}


@router.post("/master/destinations/{dest_id}/unmerge")
async def unmerge_master_destination(dest_id: str, user=Depends(MASTER)):
    """BATALKAN gabungan (batch 6): dokumen yang IKUT PINDAH saat merge (tercatat di
    `merged_moved`) dikembalikan memakai nama sumber — hanya bila destinasinya masih nama
    target (dokumen yang sudah diubah manual sesudah merge TIDAK disentuh, dilaporkan
    sebagai `skipped`). Sumber aktif kembali di selector."""
    db = get_db()
    source = await db.destinations.find_one({"id": dest_id, "deleted": {"$ne": True}}, {"_id": 0})
    if not source:
        raise HTTPException(status_code=404, detail="Destinasi tidak ditemukan")
    if not source.get("merged_into"):
        raise HTTPException(status_code=400, detail="Destinasi ini tidak sedang dalam status gabungan")
    src_name = _clean(source.get("name"))
    target = await db.destinations.find_one({"id": source["merged_into"]}, {"_id": 0})
    tgt_name = _clean((target or {}).get("name")) or _clean(source.get("merged_into_name"))
    moved = source.get("merged_moved") or {}
    restored, skipped = {"bookings": 0, "leads": 0, "quotations": 0}, 0
    for coll in ("bookings", "leads", "quotations"):
        for doc_id in (moved.get(coll) or []):
            res = await db[coll].update_one(
                {"id": doc_id, "destination": tgt_name}, {"$set": {"destination": src_name}})
            if res.modified_count:
                restored[coll] += 1
            else:
                skipped += 1
    await db.destinations.update_one({"id": dest_id}, {
        "$set": {"ops_active": True, "updated_at": now_iso()},
        "$unset": {"merged_into": "", "merged_into_name": "", "merged_moved": ""}})
    await record(db, actor=user, action="update", entity_type="destination", entity_id=dest_id,
                 summary=f"BATAL GABUNG destinasi: '{src_name}' dipulihkan dari '{tgt_name}' "
                         f"· {sum(restored.values())} dokumen kembali, {skipped} dilewati")
    return {"unmerged": True, "source": src_name, "from_target": tgt_name,
            "restored": restored, "skipped": skipped}


# ---------- KOTA (master utk customers.city & partners.city — INV-REF-02 batch 5) ----------
@router.get("/cities")
async def list_cities(user=Depends(CITIES)):
    """Utk selector form pelanggan/mitra (hanya yang AKTIF). Kelola: GET /master/cities."""
    rows = await get_db().cities.find(
        {"deleted": {"$ne": True}, "active": {"$ne": False}},
        {"_id": 0}).sort("name", 1).to_list(500)
    return safe_doc(rows)


@router.post("/cities")
async def create_city(body: CityCreate, user=Depends(CITIES)):
    """Quick-add master kota. Idempotent: nama yang sudah ada dikembalikan, bukan duplikat."""
    db = get_db()
    name = body.name.strip()
    async for c in db.cities.find({"deleted": {"$ne": True}}, {"_id": 0}):
        if _clean(c.get("name")).lower() == name.lower():
            if c.get("active") is False:
                await db.cities.update_one({"id": c["id"]}, {"$set": {"active": True}})
                c["active"] = True
            return safe_doc(c)
    doc = {"id": new_id("cty"), "name": name, "active": True, "created_at": now_iso()}
    await db.cities.insert_one(dict(doc))
    await record(db, actor=user, action="create", entity_type="city",
                 entity_id=doc["id"], summary=f"Kota baru: {name}")
    return doc


@router.get("/master/cities")
async def master_cities(user=Depends(MASTER)):
    """Kelola master kota: semua baris (termasuk nonaktif) + pemakaian pelanggan/mitra."""
    db = get_db()
    rows = await db.cities.find({"deleted": {"$ne": True}}, {"_id": 0}).sort("name", 1).to_list(500)
    out = []
    for r in rows:
        nm = r.get("name")
        used_c = await db.customers.count_documents({"city": nm})
        used_p = await db.partners.count_documents({"city": nm})
        used_w = await db.workshops.count_documents({"city": nm})
        out.append({**r, "active": r.get("active") is not False,
                    "used_by_customers": used_c, "used_by_partners": used_p,
                    "used_by_workshops": used_w})
    return safe_doc(out)


@router.patch("/master/cities/{city_id}")
async def update_master_city(city_id: str, body: MasterCityUpdate, user=Depends(MASTER)):
    """Rename kota (CASCADE ke customers/partners) dan/atau aktif/nonaktif."""
    db = get_db()
    city = await db.cities.find_one({"id": city_id, "deleted": {"$ne": True}}, {"_id": 0})
    if not city:
        raise HTTPException(status_code=404, detail="Kota tidak ditemukan")
    updates, cascade = {}, {"customers": 0, "partners": 0, "workshops": 0}
    new_name = _clean(body.name) if body.name is not None else ""
    old_name = _clean(city.get("name"))
    if new_name and new_name != old_name:
        if await _name_taken(db.cities, new_name, exclude_id=city_id):
            raise HTTPException(status_code=400, detail=f"Nama '{new_name}' sudah dipakai kota lain")
        updates["name"] = new_name
        for coll in ("customers", "partners", "workshops"):
            res = await db[coll].update_many({"city": old_name}, {"$set": {"city": new_name}})
            cascade[coll] = res.modified_count
    if body.active is not None:
        updates["active"] = bool(body.active)
    if not updates:
        return {**city, "cascade": cascade}
    await db.cities.update_one({"id": city_id}, {"$set": updates})
    await record(db, actor=user, action="update", entity_type="city", entity_id=city_id,
                 summary=f"Master kota: {old_name} → {updates.get('name', old_name)}"
                         f"{' (nonaktif)' if updates.get('active') is False else ''}"
                         f" · cascade {sum(cascade.values())} dokumen")
    return {**city, **updates, "cascade": cascade}


# ---------- EKSPOR EXCEL (batch 5): seluruh master utk tim ops ----------
@router.get("/master/export")
async def master_export(user=Depends(MASTER)):
    """Unduh seluruh master (titik jemput, destinasi, kota) sebagai satu berkas Excel."""
    from openpyxl import Workbook
    db = get_db()
    wb = Workbook()

    ws = wb.active
    ws.title = "Titik Jemput"
    ws.append(["Nama", "Aktif", "Dipakai (booking)"])
    async for r in db.pickup_points.find({"deleted": {"$ne": True}}, {"_id": 0}).sort("name", 1):
        used = await db.bookings.count_documents({"origin": r.get("name")})
        ws.append([r.get("name"), "Ya" if r.get("active") is not False else "Tidak", used])

    ws2 = wb.create_sheet("Destinasi")
    ws2.append(["Nama", "Status Web", "Aktif (ops)", "Booking", "Lead", "Penawaran", "Digabung ke"])
    async for r in db.destinations.find({"deleted": {"$ne": True}}, {"_id": 0}).sort("name", 1):
        nm = r.get("name")
        ws2.append([nm, r.get("status") or "-",
                    "Ya" if r.get("ops_active") is not False else "Tidak",
                    await db.bookings.count_documents({"destination": nm}),
                    await db.leads.count_documents({"destination": nm}),
                    await db.quotations.count_documents({"destination": nm}),
                    r.get("merged_into_name") or "-"])

    ws3 = wb.create_sheet("Kota")
    ws3.append(["Nama", "Aktif", "Pelanggan", "Mitra", "Bengkel"])
    async for r in db.cities.find({"deleted": {"$ne": True}}, {"_id": 0}).sort("name", 1):
        nm = r.get("name")
        ws3.append([nm, "Ya" if r.get("active") is not False else "Tidak",
                    await db.customers.count_documents({"city": nm}),
                    await db.partners.count_documents({"city": nm}),
                    await db.workshops.count_documents({"city": nm})])

    for sheet in wb.worksheets:
        for col in sheet.columns:
            width = max((len(str(c.value or "")) for c in col), default=8)
            sheet.column_dimensions[col[0].column_letter].width = min(width + 3, 42)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="master-data-rahazatrans.xlsx"'})
